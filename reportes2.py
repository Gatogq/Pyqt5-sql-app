from pandas import read_excel, concat, Timedelta, ExcelFile, to_datetime, isnull, notna, isna
from openpyxl import load_workbook
from numpy import select
from re import compile
from os.path import join, dirname, abspath
import sys
from gspread import authorize
from gspread_dataframe import set_with_dataframe
from datetime import time as datetime_time
from google.oauth2.service_account import Credentials
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive

if getattr(sys,'frozen',False):
    b_dir = sys._MEIPASS
else:
    b_dir = dirname(abspath(__file__))
    
sheet_key = '1MOHpyELhZoGTxU2svDUbnyJ1cnsd08GqxTcjGtROCHc'
service_account_key = join(b_dir,'data','aut-1777-81fea8b3057a.json')

def data_import(filepath_1,filepath_2,directory_path):

    xls_1 = ExcelFile(filepath_1)
    xls_2 = ExcelFile(filepath_2)
    xls_3 = ExcelFile(directory_path)

    return list([xls_1,xls_2,xls_3])


def hourzone_conditions(row):
    if isinstance(row['Estado'],str) and ((row['Estado'].strip().lower() == 'baja california norte' or row['Estado'].strip().lower() == 'baja california sur' or row['Estado'].strip().lower() == 'sinaloa' or row['Estado'].strip().lower() == 'sonora')):
        if (row['ALZA DE CORTINA'] == '') | (isna(to_datetime(row['HORA DE APERTURA'],errors='coerce'))):
            return 'sin informacion'
        elif to_datetime(row['ALZA DE CORTINA'],errors='coerce')<= to_datetime(row['HORA DE APERTURA'],errors='coerce')+Timedelta(minutes=+65):
            return 'Temprano'
        else:
            return 'Tarde'
    elif isinstance(row['Estado'],str) and row['Estado'].strip().lower() == 'quintana roo':
        if (row['ALZA DE CORTINA'] == '') | (isna(to_datetime(row['HORA DE APERTURA'],errors='coerce'))):
            return 'sin informacion'
        elif to_datetime(row['ALZA DE CORTINA'],errors='coerce')<= to_datetime(row['HORA DE APERTURA'],errors='coerce')-Timedelta(minutes=55):
            return 'Temprano'
        else:
            return 'Tarde'
    else:
        if (row['ALZA DE CORTINA'] == '') | (isna(to_datetime(row['HORA DE APERTURA'],errors='coerce'))):
            return 'sin informacion'
        elif to_datetime(row['ALZA DE CORTINA'],errors='coerce')<= to_datetime(row['HORA DE APERTURA'],errors='coerce')+Timedelta(minutes=5):
            return 'Temprano'
        else:
            return 'Tarde'
   
    
def append_sheets(excelfiles_list: list):
    dataframes=[]
    for excel_workbook in excelfiles_list[:2]:
        workbook = load_workbook(excel_workbook,data_only=True)
        for sheet_name in excel_workbook.sheet_names:
            df=read_excel(excel_workbook,sheet_name,skiprows=2)
            df = df.loc[:, ~df.columns.str.startswith('Unnamed')]
            if not df.empty and not df.columns.empty:
                last_col = df.columns[-1]
                if last_col.lower().startswith('o'): 
                    df = df.drop(last_col, axis=1)
            df = df.drop(columns=df.filter(regex=compile(r'\.\d+$'), axis=1).columns)
            target = 'OBSERVACIONES'
            df = df.loc[:, ~df.columns.str.strip().str.upper().isin([target.strip().upper()])]
            try:
                sheet = workbook[sheet_name]
                start_column = 2
                found_title = False
                column_index = start_column

                while not found_title:
                    cell_value = sheet.cell(row=2,column=column_index).value
                    if cell_value is not None:
                        title = cell_value
                        found_title = True
                    else:
                        column_index += 1
                        if column_index>sheet.max_column:
                            break
            except Exception as e:
                title = None
                  
            df['NOMBRE'] = title
            df['NOMBRE']= df['NOMBRE'].astype(str)
            df['PV'] = df['NOMBRE'].str.split().str[0]
            df['PV'] = df['PV'].str.strip()
            print(df)
            if not df.empty:
                
                df.columns = ['FECHA', 'ALZA DE CORTINA', 'BAJA DE CORTINA', 'NOMBRE','PV']     
                dataframes.append(df)
            else:
                pass
    concatdf = concat(dataframes,ignore_index=True)
    directory = read_excel(excelfiles_list[2],sheet_name='PVF´S')
    directory['Ceco/Clave  Hanna'] = directory['Ceco/Clave  Hanna'].str.strip()
    directory = directory[['Ceco/Clave  Hanna','Zona','Estado','Jefe Regional (en caso que no se tenga respusta con el cordinador de venta ni asistente)','Coordinador de venta','Horarios de servicio Lunes a Viernes','Horarios de servicio Sábados','Horarios de servicio Domingo']]
    fdf = concatdf.merge(directory, how='left', left_on ='PV', right_on='Ceco/Clave  Hanna')

    for excel_workbook in excelfiles_list:
        excel_workbook.close
    fdf.dropna(subset=['FECHA'], inplace=True)


    fdf['FECHA'] = fdf['FECHA'].dt.date.astype(str)
    fdf['ALZA DE CORTINA'] = fdf['ALZA DE CORTINA'].apply(lambda x: '' if isnull(x) else x)
    fdf['ALZA DE CORTINA'] = fdf['ALZA DE CORTINA'].astype(str)
    fdf['TIMESTAMP'] = to_datetime(fdf['FECHA']+ " " + fdf['ALZA DE CORTINA'],errors='coerce')
    fdf['DIA'] = to_datetime(fdf['FECHA']).dt.dayofweek

    conditions = [
    fdf['DIA']<5,
    fdf['DIA']==5,
    fdf['DIA']==6,
    ]

    choices = [
    fdf['Horarios de servicio Lunes a Viernes'].str.split().str[0],
    fdf['Horarios de servicio Sábados'].str.split().str[0],
    fdf['Horarios de servicio Domingo'].str.split().str[0]
    ]

    fdf['HORA DE APERTURA'] = select(conditions,choices)
    fdf['HORA DE APERTURA'] = fdf['HORA DE APERTURA'].str.replace('.',':')
    fdf['DIFERENCIA'] = (to_datetime(fdf['ALZA DE CORTINA'],errors='coerce')-to_datetime(fdf['HORA DE APERTURA'],errors='coerce'))
    fdf['APERTURA'] = to_datetime(fdf['HORA DE APERTURA'],errors='coerce').dt.time
    fdf['DIFERENCIA'] = fdf['DIFERENCIA'].apply(
    lambda x: f"{int(x.seconds // 3600):02}:{int((x.seconds // 60) % 60):02}:{int(x.seconds % 60):02}"
    if notna(x) and x.days >= 0
    else f"-{int(abs(x).seconds // 3600):02}:{int((abs(x).seconds // 60) % 60):02}:{int(abs(x).seconds % 60):02}"
    if notna(x)
    else ""
    )

    fdf['ESTATUS'] = fdf.apply(hourzone_conditions,axis=1)
    fdf.drop(columns=[
        'APERTURA',
        'Horarios de servicio Lunes a Viernes',
        'Horarios de servicio Domingo',
        'Horarios de servicio Sábados',
        'Ceco/Clave  Hanna',
        'BAJA DE CORTINA',
        'TIMESTAMP'
        ],
        inplace=True)
    fdf['HORA DE APERTURA'] = to_datetime(fdf['HORA DE APERTURA'],errors='coerce').dt.time
    fdf.rename(columns={
       'Jefe Regional (en caso que no se tenga respusta con el cordinador de venta ni asistente)':'JEFE REGIONAL',
        'Coordinador de venta':'COORDINADOR',
        'Zona':'ZONA',
        'Estado':'ESTADO'
        },inplace=True)
    fdf = fdf.reindex(columns=[
            'PV',
            'NOMBRE',
            'ZONA',
            'ESTADO',
            'JEFE REGIONAL',
            'COORDINADOR',
            'FECHA',
            'DIA',
            'ALZA DE CORTINA',
            'HORA DE APERTURA',
            'DIFERENCIA',
            'ESTATUS'
            ])
    days = {
    0: 'Lunes',
    1: 'Martes',
    2: 'Miercoles',
    3: 'Jueves',
    4: 'Viernes',
    5: 'Sabado',
    6: 'Domingo'
    }
    fdf['DIA']= fdf['DIA'].map(days)
    fdf= fdf.dropna(subset=['FECHA'])
    fdf['HORA DE APERTURA'] = fdf['HORA DE APERTURA'].apply(lambda x: '' if isnull(x) else x)
    fdf['ALZA DE CORTINA'] = fdf['ALZA DE CORTINA'].astype(str)
    fdf['HORA DE APERTURA'] = fdf['HORA DE APERTURA'].astype(str)
    fdf['DIFERENCIA'] = fdf['DIFERENCIA'].astype(str)

    return fdf

def dataframe_to_google_sheets(df,service_account_key,sheet_key,sheet_name):
    
    scopes = ['https://www.googleapis.com/auth/spreadsheets',
              'https://www.googleapis.com/auth/drive']
    credentials = Credentials.from_service_account_file(service_account_key, scopes=scopes)
    rec=len(df)
    gc = authorize(credentials)
    gauth = GoogleAuth()
    drive = GoogleDrive(gauth)
    gs = gc.open_by_key(sheet_key)
    worksheet = gs.worksheet(sheet_name)
    values = df.values.tolist()
    gs.values_append(sheet_name, {'valueInputOption': 'USER_ENTERED'}, {'values': values})

    return print(f'Succesfully written {rec} records into sheet with key {sheet_key}, See at: https://docs.google.com/spreadsheets/d/{sheet_key}')



def reporte_retardos(filepath_1,filepath_2,directory_path,sheet_name,month):
    excelfiles_list = data_import(filepath_1,filepath_2,directory_path)
    report = append_sheets(excelfiles_list)
    report = report.fillna('')
    report['MES'] = to_datetime(report['FECHA'],errors='coerce').dt.month
    report = report[report['MES'] == month]
    dataframe_to_google_sheets(report,service_account_key,sheet_key,sheet_name)
    
    return len(report), sheet_key

if __name__ == "__main__":
    #pass
    reporte_retardos(r"C:\Users\gabriel.tuyub\Downloads\APERTURA Y CIERRE PVF CIUDAD DE MEXICO Y AREA METROPOLITANA AGOSTO 2023.xlsx",r"C:\Users\gabriel.tuyub\Downloads\APERTURA Y CIERRE PVF FORANEAS AGOSTO 2023.xlsx",r"C:\Users\gabriel.tuyub\Downloads\DIRECTORIO SUCURSALES AGOSTO,2023 SUPRACARE.xlsx",'Reporte',8)


    
    



